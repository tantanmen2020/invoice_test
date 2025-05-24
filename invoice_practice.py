import openpyxl as excel


book = excel.load_workbook("original_data.xlsx")

totals = {}

new_book = excel.Workbook()
new_sheet = new_book.active
new_sheet.title = "合計"

customer_list = ["日付", "氏名", "金額"]
new_sheet.append(customer_list)

for sheet_name in book.sheetnames:
    sheet = book[sheet_name]

    date = sheet["例1"].value
    name = sheet["例2"].value
    amount = sheet["例3"].value

    if isinstance(amount, (int, float)) and name is not None:
        if name in totals:
            totals[name] += amount
        else:
            totals[name] = amount

for name, total in totals.items():
    new_sheet.append(["-", name, total])

new_book.save("合計.xlsx")

print("ファイルが完成しました！")
